from flask import Flask, render_template
from openpyxl import load_workbook

app= Flask(__name__)

@app.route('/')
def index():
    #loading worksheet
    workbook = load_workbook('test.xlsx')
    sheet = workbook.active

    #Reading data from spreadsheet
    data=[]
    for row in sheet.iter_rows(values_only= True):
        data.append(row)

    # pass the data to the template and render it
    return render_template('Dashboard.html', data=data)
if __name__ == '__main__':
    app.run(debug=True)